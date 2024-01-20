import pandas as pd  # Import pandas library for data analysis
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from cli import Style  # Import Style class from cli module
from formulas import sahm_ECD_darsad, sahm_sherkat_darsad, vres  # Import functions from formulas module
from log_time import log_the_time  # Import log_the_time function from log_time module
from config import files_path  # Import files_path dictionary from config module
import os


# transaction part
class Transaction:
    def __init__(self, one_third_one_path, terminal_work_path, cumulative_terminal_path):
        self.one_third_one_path = one_third_one_path
        self.terminal_work_path = terminal_work_path
        self.cumilative_terminal = cumulative_terminal_path

    # start reading Excel files and converting them to df
    @property
    def one_third_one_excel_file(self):
        # Read Excel file and convert it to a pandas dataframe
        df = pd.read_excel(
            self.one_third_one_path,
            sheet_name='monthly_20'
        )
        return df

    @property
    def zero_transaction_excel_file(self):
        # Read Excel file and convert it to a pandas dataframe
        df = pd.read_excel(
            self.terminal_work_path
        )
        return df

    @property
    def cumulative_excel_file(self):
        # Read Excel file and convert it to a pandas dataframe
        df = pd.read_excel(
            self.cumilative_terminal,
            sheet_name='Sheet1'
        )
        return df

    # end reading Excel files and converting them to df

    @staticmethod
    def creating_zero_terminals(zero_terminal_df):
        """
        Create a new dataframe for zero terminals.
        param zero_terminal_df: a pandas dataframe containing the data for zero terminals
        :return: a pandas dataframe with the "شماره ترمینال" column
        """
        zero_terminals = {
            "شماره ترمینال": zero_terminal_df["شماره ترمینال"]
        }
        return pd.DataFrame(zero_terminals)

    @staticmethod
    def inserting_1_31_file(one_third_one_df):
        """
        creating the column of main df
        :param one_third_one_df: DataFrame containing data from the 1st to the 31st of the month
        :return: DataFrame with relevant columns for our main DataFrame
        """
        # Create a dictionary containing relevant columns from the given DataFrame
        one_third_one_dict = {
            "شماره ترمینال": one_third_one_df["terminal"],
            "تعداد تراکنش": one_third_one_df["total_count"],
            "مبلغ تراکنش": one_third_one_df["total_amount"],
            "درآمد شاپرکی": one_third_one_df["total_income"],
            "تعداد موجودی": one_third_one_df["balance_count"],
            "درآمد موجودی": one_third_one_df["balance_income"],
            "تعداد شارژ": one_third_one_df["charge_count"],
            "درآمد شارژ": one_third_one_df["charge_income"],
            "مبلغ شارژ": one_third_one_df["charge_amount"],
            "تعداد خرید": one_third_one_df["buy_count"],
            "مبلغ خرید": one_third_one_df["buy_amount"],
            "درآمد خرید": one_third_one_df["buy_income"],
            "کمتر از 50000": one_third_one_df["lte5"],
            "بین 50000 تا 250000": one_third_one_df["lte25"],
            "بالاتر از 250000": one_third_one_df["gt25"],
            "تعداد تراکنش شاخص": one_third_one_df["indexes"],
        }
        # Convert the dictionary to a DataFrame
        return pd.DataFrame(one_third_one_dict)

    def inserting_values_part1(self, zero_terminals, one_third_one_df):
        """
        insert data from 'one_third_one_df' and 'zero_terminals' into the main DataFrame
        replace NaN values with 0
        :param zero_terminals: DataFrame containing data of terminals with zero transactions
        :param one_third_one_df: DataFrame containing data from the 1st to the 31st of the month
        :return: main DataFrame with new data
        """
        # Insert data from 'one_third_one_df' to our main DataFrame using 'inserting_1_31_file' method
        nest_1 = self.inserting_1_31_file(one_third_one_df)
        # Concatenate 'zero_terminals' DataFrame and 'nest_1'
        nest_2 = pd.concat([nest_1, zero_terminals])
        # Replace NaN values with 0
        main_df = nest_2.where(
            pd.notnull(nest_2), 0
        )
        # Return the resulting DataFrame
        return main_df

    @staticmethod
    def vlookup_and_inserting_to_main_df(main_df, cumulative_df):
        """
        insert data to main def using vlookup on cumulative df
        :param main_df: the main dataframe to which the cumulative dataframe needs to be added
        :param cumulative_df: the cumulative dataframe which will be merged with the main dataframe
        :return: the main dataframe with the vlookup data
        """

        # Set the index of the main dataframe to "شماره ترمینال"

        main_df.set_index("شماره ترمینال", inplace=True)
        duplicate_mask = main_df.index.duplicated()
        main_df = main_df[~duplicate_mask]
        # Create a dictionary with the selected columns from the cumulative dataframe and set its index to "شماره ترمینال"
        cumulative_data = {
            "شماره ترمینال": cumulative_df["شماره ترمینال"],
            "عنوان": cumulative_df["عنوان"],
            "استان": cumulative_df["استان"],
            "شهر": cumulative_df["شهر"],
            "شعب": cumulative_df["شعب"],
            "پروژه دماوند": cumulative_df["دفاتر دماوند"],
            "نوع درخواست": cumulative_df["امانی با ودیعه"],
            "مدل دستگاه": cumulative_df["مدل دستگاه"]
        }
        cumulative_df = pd.DataFrame(cumulative_data)
        cumulative_df.set_index("شماره ترمینال", inplace=True)
        cum_duplicate_mask = cumulative_df.index.duplicated()
        cumulative_df = cumulative_df[~cum_duplicate_mask]

        # Use the index of the main dataframe to get the corresponding values from the cumulative dataframe
        pd.options.mode.chained_assignment = None
        main_df["عنوان"] = main_df.index.map(cumulative_df["عنوان"])
        main_df["استان"] = main_df.index.map(cumulative_df["استان"])
        main_df["شهر"] = main_df.index.map(cumulative_df["شهر"])
        main_df["شعب"] = main_df.index.map(cumulative_df["شعب"])
        main_df["پروژه دماوند"] = main_df.index.map(cumulative_df["پروژه دماوند"])
        main_df["نوع درخواست"] = main_df.index.map(cumulative_df["نوع درخواست"])
        main_df["مدل دستگاه"] = main_df.index.map(cumulative_df["مدل دستگاه"])

        # Replace NaN values with "موجودنیست" (which means "not available")
        print("replacing nan value with not available")
        main_df = main_df.where(pd.notnull(main_df), "موجودنیست")

        return main_df

    # This function calculates several formula values based on the main_df data

    @staticmethod
    def calculate_formula_values_part1(main_df):

        # Calculate "درآمد خرید بدون ارزش افزوده" by dividing "درآمد خرید" by 1.09
        main_df["درآمد خرید بدون ارزش افزوده"] = \
            main_df["درآمد خرید"] / 1.09

        # Calculate "سهم ECD(درصد)" based on "تعداد تراکنش" using the sahm_ECD_darsad function
        main_df["سهم ECD(درصد)"] = \
            main_df["تعداد تراکنش"].apply(sahm_ECD_darsad)

        # Calculate "سهم شرکت(درصد)" based on "تعداد خرید" using the sahm_sherkat_darsad function
        main_df["سهم شرکت(درصد)"] = \
            main_df["تعداد خرید"].apply(sahm_sherkat_darsad)

        # Calculate "سهم GPGC(درصد)" by subtracting "سهم شرکت(درصد)" and "سهم ECD(درصد)" from 1
        main_df["سهم GPGC(درصد)"] = \
            1 - main_df["سهم شرکت(درصد)"] - \
            main_df["سهم ECD(درصد)"]

        # Calculate "سهم ECD(مبلغ)" based on "درآمد شاپرکی" and "سهم ECD(درصد)"
        main_df["سهم ECD(مبلغ)"] = \
            main_df["درآمد شاپرکی"] * \
            main_df["سهم ECD(درصد)"]

        # Calculate "سهم شرکت(مبلغ)" based on "درآمد خرید بدون ارزش افزوده" and "سهم شرکت(درصد)"

        main_df["سهم شرکت(مبلغ)"] = main_df["درآمد خرید بدون ارزش افزوده"] * main_df["سهم شرکت(درصد)"]

        # Calculate "سهم GPGC(مبلغ)" by subtracting "سهم ECD(مبلغ)" and "سهم شرکت(مبلغ)" from "درآمد شاپرکی"
        main_df["سهم GPGC(مبلغ)"] = \
            main_df["درآمد شاپرکی"] - \
            main_df["سهم ECD(مبلغ)"] - \
            main_df["سهم شرکت(مبلغ)"]

        return main_df

    @staticmethod
    def calculate_formula_values_part2_and_part3(main_df):
        """
        This function applies the `vres` function to the "تعداد تراکنش" column of the input DataFrame.
        It then filters the whole DataFrame by "پروژه دماوند" equal to 1.
        For each value of `i` in the range 1 to 14, it calculates the count and sum of "تعداد خرید"
        for the filtered DataFrame where "vres" is equal to `i`.
        It then saves the original DataFrame to an Excel file named "main file before formula values part3 .xlsx".
        The counts and sums are added to a copy of the original DataFrame read from the Excel file.
        The copy is indexed by "شماره ترمینال" and returned.
        :param main_df: the input DataFrame
        :return: the copy of the input DataFrame with the counts and sums added
        """
        # Apply `vres` function to "تعداد تراکنش" column
        main_df["vres"] = main_df["تعداد تراکنش"].apply(vres)

        # Filter the whole DataFrame by "پروژه دماوند" equal to 1
        print("Filtering main_df by damavand project equal to 1")
        main_df_damavand_filtered = main_df[main_df["پروژه دماوند"] == 1]

        # Calculate counts and sums for each value of `i` from 1 to 14
        all_counts = []
        all_sum = []
        for i in range(1, 15):
            all_counts.append(len(main_df_damavand_filtered[main_df_damavand_filtered.vres == i]))
            all_sum.append(main_df_damavand_filtered[main_df_damavand_filtered.vres == i]["تعداد خرید"].sum())

        # Save the original DataFrame to an Excel file
        main_df.to_excel('main file before formula values part3 .xlsx')

        # Read the saved DataFrame from the Excel file and add counts and sums columns
        main_df_with_formula_values_part3 = pd.read_excel('main file before formula values part3 .xlsx')
        os.remove('main file before formula values part3 .xlsx')
        main_df_with_formula_values_part3["counts"] = pd.Series(all_counts)
        main_df_with_formula_values_part3["sums"] = pd.Series(all_sum)
        main_df_with_formula_values_part3.set_index("شماره ترمینال", inplace=True)

        return main_df_with_formula_values_part3

    @staticmethod
    def hedayat_trarakonesh_main(sheet_1):
        # A list of company names
        company_names = [
            "آباده پارس",
            "آتی پرداز گیلانیان",
            "ارتباط دیده بان اورانوس آینده",
            "ارمغان تجارت کیان رایکا",
            "افلاک رایانه فردوس",
            "بازرگانی آیریک تجارت نصف جهان",
            "تجارت الکترونیک مینو جهان",
            "تجارت الکترونیک هونام",
            "توسعه و تجارت صفر و یک پارسی هزاره سوم",
            "داده سازان کارت پرنیان",
            "داده کاوان دوران",
            "درگاه مهر آژمان",
            "دماوند مهر ایرانیان",
            "دیاکو پردازش توسعه شهر",
            "دیاکو پردازش هلیا",
            "دیبا توان مهرگان",
            "رسام الکترونیک مبین",
            "سپهر الماس یزدان",
            "سپیدار صنعت سپهرداد",
            "شرکت آرشا تجارت ماه دیس",
            "شینا ارتباط نوین فراد کارت",
            "فن آوا کارت",
            "فناوران اسپاد شرق پوریا",
            "کیان پرداز زاگرس ",
            "گرشا تجارت هخامنش",
            "مدرن پرداخت مبین",
            "هونام رایانه افزار اکسون",
            "ویرا تجارت اوستا"
        ]
        # Prints the number of companies in the list
        print(f"we have {len(company_names)} company's")
        # Prints a message about filtering the companies
        print("filtering company's")
        # Filters the sheet_1 dataframe by companies in the company_names list
        sheet_1_filtered_by_company = sheet_1[sheet_1["شعب"].isin(company_names)]
        # Returns the filtered dataframe
        return sheet_1_filtered_by_company

    @staticmethod
    def sahm_sherkat_menha_10(x):
        """
        preparing "menha 10" file by if "gpgc" statement
        :param x: hedayat tarakonesh main
        :return: "sahm sherkat percent" hedayat tarakonesh with - 10%
        """
        # If the "sahm sherkat" percentage is 0, return 0. Otherwise, subtract 10% from the percentage.
        if x["سهم شرکت(درصد)"] == 0:
            return 0
        else:
            return x["سهم شرکت(درصد)"] - 0.10

    @staticmethod
    def sahm_gpgc_plus_10(x):
        """
        preparing "menha 10" file by if "gpgc" statement
        :param x: hedayat tarakonesh main
        :return: "GPGC percent" hedayat tarakonesh with + 10%
        """
        # If the "sahm sherkat" percentage is not 0, add 10% to the "sahm GPGC" percentage.
        # Otherwise, return the "sahm GPGC" percentage without any changes.
        if x["سهم شرکت(درصد)"] != 0:
            return x["سهم GPGC(درصد)"] + 0.10
        else:
            x["سهم GPGC(درصد)"] = x["سهم GPGC(درصد)"]
            return x["سهم GPGC(درصد)"]

    @log_the_time
    def start(self):
        # Reading excel files and creating base dataframes
        print("reading excel files")
        print("creating base dataframes")
        one_third_one_df = self.one_third_one_excel_file
        zero_transaction_df = self.zero_transaction_excel_file
        cumulative_df = self.cumulative_excel_file
        print(Style.GREEN + "dataframes created" + Style.RESET)

        # Creating transaction df and inserting values to main df
        print("creating transaction df started...")
        print("creating zero terminals df")
        zero_terminals = self.creating_zero_terminals(zero_transaction_df)
        main_df = self.inserting_values_part1(zero_terminals, one_third_one_df)
        print("doing vlookup on values and inserting them to main df")
        main_df = self.vlookup_and_inserting_to_main_df(main_df, cumulative_df)

        # Calculating formula values for main df
        print("creating formula values part 1")
        main_df = self.calculate_formula_values_part1(main_df)
        print("creating formula values part 2 & 3")
        sheet_1_main = self.calculate_formula_values_part2_and_part3(main_df)

        # Creating Sheet1 main xlsx file
        print("creating xlsx file Sheet1")
        sheet_1_main.to_excel(f" {files_path['month']} sheet1 main.xlsx")
        print(Style.GREEN + "Sheet1 main created" + Style.RESET)

        # Creating hedayat tarakonesh main and hedayat tarakonesh menha 10 xlsx files
        print("creating hedayat tarakonesh main")
        hedayat_tarakonesh_main = self.hedayat_trarakonesh_main(sheet_1_main)
        hedayat_tarakonesh_main.to_excel(f" {files_path['month']} hedayat tarakonesh main.xlsx")
        print(Style.GREEN + "hedayat tarakonesh main created" + Style.RESET)
        print("creating hedayat tarakonesh menha 10")
        hedayat_tarakonesh_main = pd.read_excel(f" {files_path['month']} hedayat tarakonesh main.xlsx")
        hedayat_tarakonesh_main["سهم شرکت(درصد)"] = hedayat_tarakonesh_main.apply(self.sahm_sherkat_menha_10, axis=1)
        hedayat_tarakonesh_main["سهم شرکت(مبلغ)"] = hedayat_tarakonesh_main["درآمد خرید بدون ارزش افزوده"] * hedayat_tarakonesh_main["سهم شرکت(درصد)"]
        hedayat_tarakonesh_main["سهم GPGC(درصد)"] = hedayat_tarakonesh_main.apply(self.sahm_gpgc_plus_10, axis=1)
        hedayat_tarakonesh_main.to_excel(f" {files_path['month']} hedayat tarakonesh menha -10.xlsx")
        print(Style.GREEN + "hedayat tarakonesh menha 10 created" + Style.RESET)

        bg_color = 'FFFF00'

        dataframe = pd.read_excel(f" {files_path['month']} hedayat tarakonesh menha -10.xlsx")

        # Write the DataFrame to the Excel file
        with pd.ExcelWriter(f" {files_path['month']} hedayat tarakonesh menha -10.xlsx", engine="openpyxl", mode="a",
                            if_sheet_exists="replace") as writer:
            dataframe.to_excel(writer, sheet_name="Sheet1", index=False)

            # Get the worksheet
            workbook = writer.book
            worksheet = workbook["Sheet1"]

            # Adjust the column widths
            for col_num, col_name in enumerate(dataframe.columns, start=1):
                cell = worksheet.cell(row=1, column=col_num + 1)
                cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                column_letter = get_column_letter(col_num)
                worksheet.column_dimensions[column_letter].width = 18

            # Save the changes to the workbook
            workbook.save(f" {files_path['month']} hedayat tarakonesh menha -10.xlsx")
        # Printing final message
        print(Style.CYAN + "Transaction Part Done" + Style.RESET)
