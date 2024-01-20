# Importing necessary library
import pandas as pd
from config import files_path
from jdatetime import datetime
from cli import Style
from log_time import log_the_time
import numpy as np
from cli import center_and_justify
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill
import os


class Transfer:
    date = datetime.strptime(files_path["date"], '%Y_%m').date()  # Parsing the date from files_path and storing it in the 'date' class attribute

    def __init__(self, hedayat_tarakonesh_path):
        self.hedayat_tarakonesh_path = hedayat_tarakonesh_path  # Initializing the 'hedayat_tarakonesh_path' instance variable with the provided path

    @property
    def hedayat_tarakonesh_file(self):
        df = pd.read_excel(
            self.hedayat_tarakonesh_path,
            sheet_name='Sheet1'
        )  # Reading an Excel file from 'hedayat_tarakonesh_path' and storing the contents in the 'df' DataFrame
        return df  # Returning the DataFrame

    def create_riz_sheet(self, hedayat_tarakonesh_df):
        riz_df = hedayat_tarakonesh_df  # Assigning 'hedayat_tarakonesh_df' to 'riz_df'
        riz_df["ماه"] = self.date  # Adding a new column 'ماه' to 'riz_df' and setting its values to the stored date
        return riz_df  # Returning the modified DataFrame

    def create_kol_sheet(self, riz_df):
        kol = pd.DataFrame({"ماه": self.date}, index=[0])  # Creating a new DataFrame 'kol' with a single row and the 'ماه' column set to the stored date

        # Counting the number of rows in 'riz_df' where 'سهم ECD(درصد)' is equal to 0.9 and assigning it to 'تعداد سهم های 90 درصد'
        kol["تعداد سهم های 90 درصد"] = len(riz_df[riz_df['سهم ECD(درصد)'] == 0.9])

        # Counting the number of rows in 'riz_df' where 'سهم ECD(درصد)' is equal to 0.5 and assigning it to 'تعداد سهم های 50 درصد'
        kol["تعداد سهم های 50 درصد "] = len(riz_df[riz_df['سهم ECD(درصد)'] == 0.5])

        # Counting the number of rows in 'riz_df' where 'سهم ECD(درصد)' is equal to 0.2 and assigning it to 'تعداد سهم های 20 درصد'
        kol["تعداد سهم های 20 درصد "] = len(riz_df[riz_df['سهم ECD(درصد)'] == 0.2])

        # Calculating the total number of shares by adding the counts of shares in each bracket
        kol["تعداد کل درصد ها"] = kol["تعداد سهم های 90 درصد"] + kol["تعداد سهم های 50 درصد "] + kol["تعداد سهم های 20 درصد "]

        # Calculating the transaction amounts for each income bracket and the total transaction amount
        kol["تراکنش رده اول"] = riz_df.loc[(riz_df['سهم ECD(درصد)'] == 0.9), 'تعداد خرید'].sum()
        kol["تراکنش رده دوم"] = riz_df.loc[(riz_df['سهم ECD(درصد)'] == 0.5), 'تعداد خرید'].sum()
        kol["تراکنش رده سوم"] = riz_df.loc[(riz_df['سهم ECD(درصد)'] == 0.2), 'تعداد خرید'].sum()
        kol["جمع تراکنش ها"] = kol["تراکنش رده اول"] + kol["تراکنش رده دوم"] + kol["تراکنش رده سوم"]

        # Calculating the income for the first income bracket by summing the values from the corresponding columns
        first_income_bracket_1 = riz_df.loc[(riz_df['سهم ECD(درصد)'] == 0.9), 'سهم شرکت(مبلغ)'].sum()
        first_income_bracket_2 = riz_df.loc[(riz_df['سهم ECD(درصد)'] == 0.9), 'سهم GPGC(مبلغ)'].sum()
        first_income_bracket = first_income_bracket_1 + first_income_bracket_2
        kol["در آمد رده اول"] = first_income_bracket

        # Calculating the income for the second income bracket by summing the values from the corresponding columns
        second_income_bracket_1 = riz_df.loc[(riz_df['سهم ECD(درصد)'] == 0.5), 'سهم شرکت(مبلغ)'].sum()
        second_income_bracket_2 = riz_df.loc[(riz_df['سهم ECD(درصد)'] == 0.5), 'سهم GPGC(مبلغ)'].sum()
        second_income_bracket = second_income_bracket_1 + second_income_bracket_2
        kol["در آمد رده دوم"] = second_income_bracket

        # Calculating the income for the third income bracket by summing the values from the corresponding columns
        third_income_bracket_1 = riz_df.loc[(riz_df['سهم ECD(درصد)'] == 0.2), 'سهم شرکت(مبلغ)'].sum()
        third_income_bracket_2 = riz_df.loc[(riz_df['سهم ECD(درصد)'] == 0.2), 'سهم GPGC(مبلغ)'].sum()
        third_income_bracket = third_income_bracket_1 + third_income_bracket_2
        kol["در آمد رده سوم"] = third_income_bracket

        # Calculating the total income by summing the incomes of all income brackets
        kol["در آمد کل"] = kol["در آمد رده اول"] + kol["در آمد رده دوم"] + kol["در آمد رده سوم"]

        # Getting the unique company names from the 'شعب' column in 'riz_df'
        company_names = list(set(riz_df['شعب']))

        # Creating a dictionary to store the shares of each company
        company_shares = {}

        # Tracking the maximum number of shares for any company
        max_num_shares = 0

        # Calculating the shares for each company
        for name in company_names:
            shares = [riz_df.loc[riz_df['شعب'] == name, 'سهم شرکت(مبلغ)'].sum()]
            company_shares[name] = shares
            max_num_shares = max(max_num_shares, len(shares))

        # Adding NaN values to the end of lists for companies that have fewer shares than the maximum number of shares
        for name in company_names:
            diff = max_num_shares - len(company_shares[name])
            if diff > 0:
                company_shares[name].extend([np.nan] * diff)

        # Creating a DataFrame 'companies_df' from the company shares dictionary
        companies_df = pd.DataFrame(company_shares)

        # Concatenating 'kol' and 'companies_df' along the columns axis
        kol = pd.concat([kol, companies_df], axis=1)

        # Calculating the total shares of all companies
        kol["سهم شرکتها"] = companies_df.values.sum()

        # Calculating the share of 'سیراف' by subtracting the total company shares from the total income
        kol["سهم سیراف"] = kol["در آمد کل"] - kol["سهم شرکتها"]

        # Calculating the share of 'دماوند' by summing the values in the 'سهم ECD(مبلغ)' column of 'riz_df'
        kol["سهم دماوند"] = riz_df["سهم ECD(مبلغ)"].sum()

        return kol  # Returning the final DataFrame

    @staticmethod
    def create_damavand_invoice(kol_df):
        # Set the maximum column width for display
        pd.options.display.max_colwidth = 1000

        # Create a DataFrame for damavand_invoice_data
        damavand_invoice_data = pd.DataFrame({
            'بازه تراکنشی': [
                'تعداد دستگاه',
                'تعداد تراکنش',
                'سهم شرکت سیراف',
                'درآمد شرکت سیراف',
                'درآمد کل',
                'تعداد کل دستگاه',
                'تعداد کل تراکنش',
                'درآمد کل شرکت سیراف'
            ],
            '0 تا 100': [
                kol_df["تعداد سهم های 90 درصد"].values[0],
                kol_df["تراکنش رده اول"].values[0],
                '10%',
                kol_df["در آمد رده اول"].values[0],
                kol_df["در آمد رده اول"].values[0] / 0.1,
                kol_df["تعداد سهم های 90 درصد"].values[0] + kol_df["تعداد سهم های 50 درصد "].values[0] +
                kol_df["تعداد سهم های 20 درصد "].values[0],
                kol_df["تراکنش رده اول"].values[0] + kol_df["تراکنش رده دوم"].values[0] +
                kol_df["تراکنش رده سوم"].values[0],
                kol_df["در آمد رده اول"].values[0] + kol_df["در آمد رده دوم"].values[0] +
                kol_df["در آمد رده سوم"].values[0]
            ],
            '100 تا 250': [
                kol_df["تعداد سهم های 50 درصد "].values[0],
                kol_df["تراکنش رده دوم"].values[0],
                '50%',
                kol_df["در آمد رده دوم"].values[0],
                kol_df["در آمد رده دوم"].values[0] / 0.5,
                np.nan,
                np.nan,
                np.nan,
            ],
            '250 به بالا': [
                kol_df["تعداد سهم های 20 درصد "].values[0],
                kol_df["تراکنش رده سوم"].values[0],
                '80%',
                kol_df["در آمد رده سوم"].values[0],
                kol_df["در آمد رده سوم"].values[0] / 0.8,
                np.nan,
                np.nan,
                np.nan,
            ]
        })

        # Apply a style to center and justify the data in damavand_invoice_data
        damavand_invoice_data = damavand_invoice_data.style.applymap(center_and_justify)

        # Format numerical columns in the DataFrame
        for col in damavand_invoice_data.columns[1:]:
            damavand_invoice_data.data[col] = damavand_invoice_data.data[col].apply(
                lambda x: '{:,.0f}'.format(float(x)) if pd.notnull(x) and isinstance(x, (int, float)) else x
            )

        # Save damavand_invoice_data to an Excel file named "damavand_invoice_data.xlsx"
        damavand_invoice_data.to_excel("damavand_invoice_data.xlsx")

        # Set background color for the header row in the Excel file
        bg_color = 'FFFF00'

        # Append damavand_invoice_data to the existing Excel file
        with pd.ExcelWriter(f"damavand_invoice_data.xlsx", engine="openpyxl", mode="a",
                            if_sheet_exists="replace") as writer:
            damavand_invoice_data.to_excel(writer, sheet_name="Sheet1", index=False)
            workbook = writer.book
            worksheet = workbook["Sheet1"]

            # Adjust the column widths and set the header row's background color
            for col_num, col_name in enumerate(damavand_invoice_data.columns, start=1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                column_letter = get_column_letter(col_num)
                worksheet.column_dimensions[column_letter].width = 18

            # Save the changes to the workbook
            workbook.save(f"damavand_invoice_data.xlsx")

    @staticmethod
    def create_riz_sherkat(riz_sheet):

        # Load the data into a DataFrame
        desired_columns = [
            'عنوان',
            'استان',
            'شهر',
            'شعب',
            'شماره ترمینال',
            'تعداد خرید',
            'کمتر از 50000',
            'بین 50000 تا 250000',
            'بالاتر از 250000',
            'سهم شرکت(مبلغ)',
            'ماه',
        ]

        df = riz_sheet.loc[:, desired_columns]
        for col in df.columns[1:]:
            if col != 'شماره ترمینال':
                df[col] = df[col].apply(
                    lambda x: '{:,.0f}'.format(float(x)) if pd.notnull(x) and isinstance(x, (int, float)) else x
                )
        # Get the unique company names
        company_names = df['شعب'].unique()
        border_style = Border(left=Side(border_style='thin'),
                              right=Side(border_style='thin'),
                              top=Side(border_style='thin'),
                              bottom=Side(border_style='thin'))

        if not os.path.exists("companies excel files"):
            os.makedirs("companies excel files")

        for name in company_names:
            # Filter the DataFrame by company name
            filtered_df = df.loc[df['شعب'] == name]

            # Create a new Excel file for the filtered DataFrame
            with pd.ExcelWriter(os.path.join("companies excel files", f'{name}.xlsx'), engine='openpyxl',
                                mode='w') as writer:
                filtered_df.to_excel(writer, sheet_name='Sheet1', index=False)
                workbook = writer.book
                worksheet = workbook.active

                # Adjust the column widths and formatting
                for col_num, col_name in enumerate(filtered_df.columns, start=1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    column_letter = get_column_letter(col_num)
                    worksheet.column_dimensions[column_letter].width = 18

                # Set the cell alignment to center and the border style
                for row_num, row in enumerate(worksheet.iter_rows(min_row=1, max_col=len(filtered_df.columns)),
                                              start=1):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.border = border_style

                # Save the changes to the workbook
                workbook.save(os.path.join("companies excel files", f'{name}.xlsx'))

    @log_the_time
    def start(self):
        print("reading hedayat tarakonesh file")
        hedayat_tarakonesh_df = self.hedayat_tarakonesh_file
        print("creating transfer riz sheet")
        riz_df = self.create_riz_sheet(hedayat_tarakonesh_df)
        print(Style.GREEN + "riz sheet created" + Style.RESET)
        riz_df.to_excel("ریز.xlsx")
        print("creating kol sheet")
        kol_df = self.create_kol_sheet(riz_df)
        print(Style.GREEN + "kol sheet created" + Style.RESET)
        kol_df.to_excel("کل.xlsx")
        print("creating sorat hesab damavand")
        self.create_damavand_invoice(kol_df)
        print(Style.GREEN + "damavand sorat hesab sheet created" + Style.RESET)
        print("creating companies riz")
        self.create_riz_sherkat(riz_df)
        print(Style.GREEN + "companies riz created" + Style.RESET)
        print(Style.CYAN + "Transfer Part Done" + Style.RESET)
